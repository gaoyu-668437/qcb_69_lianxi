

#读取Excel测试数据
from openpyxl import load_workbook


def read_information(file_name,sheet_name):
    wb = load_workbook(file_name)  # 打开加载工作簿
    sheet = wb[sheet_name]  # 定位表单
    all_case = []
    for i in range(2,sheet.max_row+1):
        case = []
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row =i ,column = j).value)
        all_case.append(case)
    return all_case


def write_data(file_name,sheet_name,row,column,value):
    wb = load_workbook(file_name)
    wb[sheet_name].cell(row=row, column=column).value = value
    wb.save(file_name)

if __name__ == '__main__':
    print(read_information('data_information.xlsx','recharge'))







